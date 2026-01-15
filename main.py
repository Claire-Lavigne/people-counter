import cv2
import numpy as np
import sys
import time
from PyQt5 import QtWidgets, QtGui, QtCore
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

import os
from ultralytics import YOLO
import csv
import glob

# Fonction pour charger les ressources compatible PyInstaller et exécution directe
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# Load YOLOv8 model (CPU only for compatibility)
yolo_model = YOLO('yolov8n.pt')  # Utilise le modèle nano pour la vitesse, ou 'yolov8s.pt' pour plus de précision
print('Using CPU for YOLOv8 (ultralytics)')

# --- PyQt5 GUI ---
class VideoThread(QtCore.QThread):
    change_pixmap_signal = QtCore.pyqtSignal(np.ndarray, int)
    def __init__(self, source):
        super().__init__()
        self.source = source
        self._run_flag = True
        self.history = []  # (timestamp, count)
    def run(self):
        cap = cv2.VideoCapture(self.source)
        while self._run_flag:
            ret, frame = cap.read()
            if not ret:
                break
            # YOLOv8 attend aussi du BGR (OpenCV default)
            results = yolo_model(frame)
            # Get person detections (class 0 in COCO)
            person_count = 0
            boxes = []
            for r in results:
                for box in r.boxes:
                    cls = int(box.cls[0])
                    if cls == 0:  # class 0 = person
                        person_count += 1
                        x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                        boxes.append((x1, y1, x2, y2))
                        cv2.rectangle(frame, (x1, y1), (x2, y2), (0,255,0), 2)
                        cv2.putText(frame, 'person', (x1, y1-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,0), 2)
            self.history.append((time.time(), person_count))
            for i in boxes:
                x1, y1, x2, y2 = i
                label = 'person'
                color = (0, 255, 0)
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                cv2.putText(frame, label, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
            cv2.putText(frame, f'People Count: {person_count}', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)
            self.change_pixmap_signal.emit(frame, person_count)
            if not self._run_flag:
                break
        cap.release()
    def stop(self):
        self._run_flag = False
        self.wait()
    def get_history(self):
        return self.history

class MainWindow(QtWidgets.QWidget):
    def save_count_to_csv(self, camera_id, timestamp, count):
        # Save to a CSV file per camera
        fname = f"counts_{camera_id.replace(':', '_').replace('/', '_')}.csv"
        with open(fname, 'a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([timestamp, count])

    def load_all_counts(self):
        # Load all CSVs into a dict: {camera_id: [(timestamp, count), ...]}
        data = {}
        for fname in glob.glob('counts_*.csv'):
            camera_id = fname[7:-4].replace('_', ':', 2).replace('_', '/', 1)
            with open(fname, 'r') as f:
                reader = csv.reader(f)
                data[camera_id] = [(float(row[0]), int(row[1])) for row in reader]
        return data

    def aggregate_counts(self, level):
        import datetime
        camera_id = self.camera_selector.currentText() if hasattr(self, 'camera_selector') else self.get_current_camera_id()
        data = self.all_counts.get(camera_id, [])
        if not data:
            return [], []
        dt_counts = [(datetime.datetime.fromtimestamp(ts), v) for ts, v in data]
        if level == 'raw':
            times = [ts for ts, _ in data]
            values = [v for _, v in data]
            return times, values
        buckets = {}
        for dt, v in dt_counts:
            if level == 'minute':
                key = dt.replace(second=0, microsecond=0)
            elif level == 'hour':
                key = dt.replace(minute=0, second=0, microsecond=0)
            elif level == 'day':
                key = dt.replace(hour=0, minute=0, second=0, microsecond=0)
            elif level == 'month':
                key = dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            elif level == 'year':
                key = dt.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                key = dt
            if key not in buckets:
                buckets[key] = []
            buckets[key].append(v)
        times = []
        values = []
        for k in sorted(buckets):
            times.append(k)
            values.append(sum(buckets[k]) / len(buckets[k]))
        return times, values

    def on_agg_changed(self, idx):
        level = self.agg_combo.currentData()
        times, values = self.aggregate_counts(level)
        self.ax.clear()
        if level == 'raw':
            self.ax.plot(self.plot_times, self.plot_data, color='blue')
            self.ax.set_xlabel('Temps (s)')
        else:
            self.ax.plot(times, values, color='blue')
            self.ax.set_xlabel(level.capitalize())
        self.ax.set_title('Historique du comptage')
        self.ax.set_ylabel('Personnes')
        self.canvas.draw()

    def export_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        import tempfile
        import os
        # Sauvegarder le graphique comme image temporaire
        tmp_img = tempfile.mktemp(suffix='.png')
        self.figure.savefig(tmp_img)
        # Créer le PDF
        pdf_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Enregistrer le PDF', '', 'PDF Files (*.pdf)')
        if not pdf_path:
            os.remove(tmp_img)
            return
        c = canvas.Canvas(pdf_path, pagesize=A4)
        c.setFont('Helvetica', 14)
        c.drawString(50, 800, 'Rapport de comptage de personnes')
        c.setFont('Helvetica', 10)
        c.drawString(50, 780, self.stats_label.text())
        c.drawImage(tmp_img, 50, 400, width=500, height=300)
        c.setFont('Helvetica', 10)
        c.drawString(50, 380, 'Historique (temps en secondes, personnes):')
        y = 360
        # Remplacer la boucle d'export par l'agrégation sélectionnée
        level = self.agg_combo.currentData()
        times, values = self.aggregate_counts(level)
        y = 360
        import datetime
        for t, v in zip(times, values):
            if level == 'raw':
                label = f'{t:.1f} s'
            else:
                label = str(t)
            c.drawString(50, y, f'{label} : {v}')
            y -= 12
            if y < 50:
                c.showPage()
                y = 800
        c.save()
        os.remove(tmp_img)
        QtWidgets.QMessageBox.information(self, 'Export PDF', 'Export PDF terminé !')
    def export_excel(self):
        import openpyxl
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font
        import os
        # Demander où sauvegarder
        xlsx_path, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Enregistrer Excel', '', 'Excel Files (*.xlsx)')
        if not xlsx_path:
            return
        level = self.agg_combo.currentData()
        times, values = self.aggregate_counts(level)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Comptage'
        ws['A1'] = 'Temps'
        ws['B1'] = 'Personnes'
        ws['A1'].font = ws['B1'].font = Font(bold=True)
        for i, (t, v) in enumerate(zip(times, values), start=2):
            ws[f'A{i}'] = str(t) if level != 'raw' else t
            ws[f'B{i}'] = v
        # Ajuster largeur colonnes
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 20
        try:
            wb.save(xlsx_path)
            QtWidgets.QMessageBox.information(self, 'Export Excel', 'Export Excel terminé !')
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Erreur', f'Erreur lors de l\'export : {e}')

    def start_video(self):
        # Démarre la capture vidéo selon la source sélectionnée
        if self.combo.currentIndex() == self.combo.count() - 1:
            url = self.url_input.text().strip()
            if not url:
                QtWidgets.QMessageBox.warning(self, 'Erreur', 'Veuillez saisir une URL de caméra réseau valide.')
                return
            source = url
        else:
            cam_indices = [int(self.combo.itemText(i).split()[-1]) for i in range(self.combo.count()-1)]
            source = cam_indices[self.combo.currentIndex()]
        print(f"[INFO] Trying to open video source: {source}")
        # Test if the camera can be opened
        test_cap = cv2.VideoCapture(source)
        if not test_cap.isOpened():
            QtWidgets.QMessageBox.critical(self, 'Erreur', f'Impossible d\'ouvrir la caméra locale (index/source: {source}).\nVérifiez qu\'elle n\'est pas utilisée par une autre application et qu\'elle est bien connectée.')
            test_cap.release()
            return
        test_cap.release()
        self.thread = VideoThread(source)
        self.thread.change_pixmap_signal.connect(self.update_image)
        self.thread.start()
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.start_time = time.time()
        self.plot_data = []
        self.plot_times = []
        self.ax.clear()
        self.ax.set_title('Historique du comptage')
        self.ax.set_xlabel('Temps (s)')
        self.ax.set_ylabel('Personnes')
        self.canvas.draw()
    def update_start_button_state(self):
        # Active le bouton Démarrer seulement si une source valide est sélectionnée
        if self.combo.currentIndex() == self.combo.count() - 1:
            # Caméra réseau : vérifier que l’URL n’est pas vide
            self.btn_start.setEnabled(bool(self.url_input.text().strip()))
        else:
            self.btn_start.setEnabled(True)
    def __init__(self):
        super().__init__()
        self.setWindowTitle('People Counter')
        self.setGeometry(100, 100, 1200, 700)
        # Main layout: horizontal split
        main_layout = QtWidgets.QHBoxLayout(self)
        # Left: video and plot
        left_panel = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_panel)
        self.image_label = QtWidgets.QLabel()
        self.image_label.setFixedSize(900, 640)
        left_layout.addWidget(self.image_label)
        self.figure = Figure(figsize=(6, 2.5))
        self.canvas = FigureCanvas(self.figure)
        self.ax = self.figure.add_subplot(111)
        self.ax.set_title('Historique du comptage')
        self.ax.set_xlabel('Temps (s)')
        self.ax.set_ylabel('Personnes')
        left_layout.addWidget(self.canvas)
        main_layout.addWidget(left_panel, 1)
        # Right: controls
        right_panel = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_panel)
        right_layout.setAlignment(QtCore.Qt.AlignTop)
        # Camera group
        self.camera_group = QtWidgets.QGroupBox('Source vidéo')
        camera_layout = QtWidgets.QVBoxLayout(self.camera_group)
        self.combo = QtWidgets.QComboBox()
        # Détection automatique des webcams disponibles
        def detect_cameras(max_test=5):
            import cv2
            available = []
            for i in range(max_test):
                cap = cv2.VideoCapture(i)
                if cap is not None and cap.isOpened():
                    available.append(i)
                    cap.release()
            return available

        cam_indices = detect_cameras()
        cam_labels = [f'Webcam locale {i}' for i in cam_indices]
        self.combo.addItems(cam_labels + ['Caméra réseau (URL)'])
        self.combo.setToolTip('Sélectionnez une webcam locale détectée ou choisissez "Caméra réseau (URL)" pour saisir un flux réseau.')
        camera_layout.addWidget(self.combo)
        self.url_input = QtWidgets.QLineEdit()
        self.url_input.setPlaceholderText('rtsp://... ou http://...')
        self.url_input.setVisible(False)
        self.url_input.setToolTip('Saisissez ici l’URL du flux réseau (RTSP/HTTP).')
        camera_layout.addWidget(self.url_input)
        self.url_help = QtWidgets.QLabel('Exemple : rtsp://192.168.1.10/stream')
        self.url_help.setStyleSheet('color: gray; font-size: 9pt;')
        self.url_help.setVisible(False)
        camera_layout.addWidget(self.url_help)
        self.combo.currentIndexChanged.connect(self.toggle_url_input)
        self.url_input.textChanged.connect(self.update_start_button_state)
        self.combo.currentIndexChanged.connect(self.update_start_button_state)
        right_layout.addWidget(self.camera_group)
        # Start/Stop buttons
        self.btn_start = QtWidgets.QPushButton('Démarrer')
        self.btn_start.clicked.connect(self.start_video)
        self.btn_stop = QtWidgets.QPushButton('Arrêter')
        self.btn_stop.setEnabled(False)
        self.btn_stop.clicked.connect(self.stop_video)
        btn_layout = QtWidgets.QHBoxLayout()
        btn_layout.addWidget(self.btn_start)
        btn_layout.addWidget(self.btn_stop)
        right_layout.addLayout(btn_layout)
        # Aggregation dropdown
        self.agg_combo = QtWidgets.QComboBox()
        self.agg_combo.addItem('Brut (chaque mesure)', 'raw')
        self.agg_combo.addItem('Par minute', 'minute')
        self.agg_combo.addItem('Par heure', 'hour')
        self.agg_combo.addItem('Par jour', 'day')
        self.agg_combo.addItem('Par mois', 'month')
        self.agg_combo.addItem('Par année', 'year')
        self.agg_combo.currentIndexChanged.connect(self.on_agg_changed)
        right_layout.addWidget(self.agg_combo)
        # Export buttons
        self.btn_export = QtWidgets.QPushButton('Exporter PDF')
        self.btn_export.clicked.connect(self.export_pdf)
        self.btn_export_excel = QtWidgets.QPushButton('Exporter Excel')
        self.btn_export_excel.clicked.connect(self.export_excel)
        right_layout.addWidget(self.btn_export)
        right_layout.addWidget(self.btn_export_excel)
        # Stats label (optional, for non-overlay display)
        self.stats_label = QtWidgets.QLabel('Moyenne: -   Min: -   Max: -')
        right_layout.addWidget(self.stats_label)
        # Add zoom controls to right panel
        self.btn_zoom_in = QtWidgets.QPushButton('Zoom +')
        self.btn_zoom_in.clicked.connect(self.zoom_in)
        self.btn_zoom_out = QtWidgets.QPushButton('Zoom -')
        self.btn_zoom_out.clicked.connect(self.zoom_out)
        self.btn_zoom_in.setEnabled(False)
        self.btn_zoom_out.setEnabled(False)
        right_layout.addWidget(self.btn_zoom_in)
        right_layout.addWidget(self.btn_zoom_out)
        right_layout.addStretch(1)
        main_layout.addWidget(right_panel, 0)
        # --- Fin de l'initialisation de l'interface ---
        # Connexions supplémentaires
        self.thread = None
        self.start_time = None
        self.plot_data = []
        self.plot_times = []
        # Add camera selector for dashboard
        self.all_counts = self.load_all_counts()
        self.camera_selector = QtWidgets.QComboBox()
        for cam in self.all_counts:
            self.camera_selector.addItem(cam)
        self.camera_selector.currentIndexChanged.connect(self.on_camera_selected)
        right_layout.addWidget(self.camera_selector)

    def toggle_url_input(self, idx):
    # Affiche le champ URL uniquement si "Caméra réseau (URL)" est sélectionné
        if idx == self.combo.count() - 1:
            self.url_input.setVisible(True)
            self.url_help.setVisible(True)
        else:
            self.url_input.setVisible(False)
            self.url_help.setVisible(False)
    def stop_video(self):
        if self.thread:
            self.thread.stop()
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
    def update_image(self, cv_img, person_count):
        qt_img = self.convert_cv_qt(cv_img)
        # Overlay stats on the video frame
        import cv2
        import numpy as np
        # Convert QPixmap back to OpenCV image for overlay
        frame = cv_img.copy()
        # Prepare stats text
        # Ne rien afficher pour les stats (moyenne, min, max)
        if hasattr(self, 'stats_label'):
            self.stats_label.setText("")
        # Draw people count
        cv2.putText(frame, f'People Count: {person_count}', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0,0,255), 2)
        # Ne pas afficher les stats sur la vidéo
        # Convert back to QPixmap
        rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(900, 640, QtCore.Qt.KeepAspectRatio)
        self.image_label.setPixmap(QtGui.QPixmap.fromImage(p))
        # Update plot
        if self.start_time is not None:
            t = time.time() - self.start_time
            self.plot_times.append(t)
            self.plot_data.append(person_count)
            self.on_agg_changed(self.agg_combo.currentIndex())
        # Save to CSV (persistent log)
        camera_id = self.get_current_camera_id()
        if self.start_time is not None:
            t = time.time() - self.start_time
            self.save_count_to_csv(camera_id, time.time(), person_count)

    def get_current_camera_id(self):
        # Use URL for network, index for local
        if self.combo.currentIndex() == self.combo.count() - 1:
            return self.url_input.text().strip()
        else:
            return f"local_{self.combo.currentIndex()}"
    def convert_cv_qt(self, cv_img):
        rgb_image = cv2.cvtColor(cv_img, cv2.COLOR_BGR2RGB)
        h, w, ch = rgb_image.shape
        bytes_per_line = ch * w
        convert_to_Qt_format = QtGui.QImage(rgb_image.data, w, h, bytes_per_line, QtGui.QImage.Format_RGB888)
        p = convert_to_Qt_format.scaled(900, 640, QtCore.Qt.KeepAspectRatio)
        return QtGui.QPixmap.fromImage(p)

    def zoom_in(self):
        url = self.url_input.text().strip()
        if not url:
            QtWidgets.QMessageBox.warning(self, 'Erreur', 'Aucune URL de caméra réseau spécifiée.')
            return
        try:
            from onvif import ONVIFCamera
            import urllib.parse
            parsed = urllib.parse.urlparse(url)
            host = parsed.hostname
            port = parsed.port or 80
            # Prompt for credentials if needed
            user, ok1 = QtWidgets.QInputDialog.getText(self, 'ONVIF Login', 'Utilisateur:', QtWidgets.QLineEdit.Normal, 'admin')
            if not ok1:
                return
            pwd, ok2 = QtWidgets.QInputDialog.getText(self, 'ONVIF Login', 'Mot de passe:', QtWidgets.QLineEdit.Password, '')
            if not ok2:
                return
            cam = ONVIFCamera(host, port, user, pwd)
            media = cam.create_media_service()
            ptz = cam.create_ptz_service()
            profiles = media.GetProfiles()
            token = profiles[0].token
            req = ptz.create_type('ContinuousMove')
            req.ProfileToken = token
            req.Velocity = {'Zoom': {'x': 0.5}}  # Zoom in
            ptz.ContinuousMove(req)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Erreur', f'Zoom IN ONVIF échoué : {e}')

    def zoom_out(self):
        url = self.url_input.text().strip()
        if not url:
            QtWidgets.QMessageBox.warning(self, 'Erreur', 'Aucune URL de caméra réseau spécifiée.')
            return
        try:
            from onvif import ONVIFCamera
            import urllib.parse
            parsed = urllib.parse.urlparse(url)
            host = parsed.hostname
            port = parsed.port or 80
            user, ok1 = QtWidgets.QInputDialog.getText(self, 'ONVIF Login', 'Utilisateur:', QtWidgets.QLineEdit.Normal, 'admin')
            if not ok1:
                return
            pwd, ok2 = QtWidgets.QInputDialog.getText(self, 'ONVIF Login', 'Mot de passe:', QtWidgets.QLineEdit.Password, '')
            if not ok2:
                return
            cam = ONVIFCamera(host, port, user, pwd)
            media = cam.create_media_service()
            ptz = cam.create_ptz_service()
            profiles = media.GetProfiles()
            token = profiles[0].token
            req = ptz.create_type('ContinuousMove')
            req.ProfileToken = token
            req.Velocity = {'Zoom': {'x': -0.5}}  # Zoom out
            ptz.ContinuousMove(req)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Erreur', f'Zoom OUT ONVIF échoué : {e}')

    def on_camera_selected(self, idx):
        self.on_agg_changed(self.agg_combo.currentIndex())

def main():
    app = QtWidgets.QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
